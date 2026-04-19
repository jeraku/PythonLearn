"""
Fridge Grocery Manager - Data Layer
Handles all Excel read/write operations for grocery items.
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, date
import os
import json
import sys

EXCEL_FILE = "fridge_groceries.xlsx"
SHEET_NAME = "Groceries"

COLUMNS = ["ID", "Item", "Category", "Quantity", "Unit", "Expiry Date", "Added Date", "Notes"]

CATEGORIES = ["Dairy", "Meat & Fish", "Vegetables", "Fruits", "Beverages", "Condiments", "Leftovers", "Other"]

SAMPLE_DATA = [
    [1, "Whole Milk", "Dairy", 2, "litre", "2026-04-22", "2026-04-15", "Semi-skimmed"],
    [2, "Cheddar Cheese", "Dairy", 300, "g", "2026-04-28", "2026-04-14", ""],
    [3, "Chicken Breast", "Meat & Fish", 500, "g", "2026-04-20", "2026-04-18", "Marinated"],
    [4, "Broccoli", "Vegetables", 1, "head", "2026-04-21", "2026-04-17", ""],
    [5, "Greek Yogurt", "Dairy", 4, "pots", "2026-04-25", "2026-04-16", "Strawberry"],
    [6, "Orange Juice", "Beverages", 1, "litre", "2026-04-23", "2026-04-15", "Fresh squeezed"],
    [7, "Salmon Fillet", "Meat & Fish", 400, "g", "2026-04-19", "2026-04-18", ""],
    [8, "Spinach", "Vegetables", 200, "g", "2026-04-20", "2026-04-17", "Baby spinach"],
    [9, "Butter", "Dairy", 250, "g", "2026-05-10", "2026-04-10", "Salted"],
    [10, "Leftover Pasta", "Leftovers", 1, "container", "2026-04-20", "2026-04-18", "Bolognese"],
]


def create_excel(filepath=EXCEL_FILE):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    header_fill = PatternFill("solid", start_color="1B4332")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    alt_fill = PatternFill("solid", start_color="F0FFF4")
    border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row_idx, row_data in enumerate(SAMPLE_DATA, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(vertical="center")
            cell.border = border
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    col_widths = [6, 22, 15, 10, 10, 14, 14, 25]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    wb.save(filepath)
    return True


def load_data(filepath=EXCEL_FILE):
    if not os.path.exists(filepath):
        create_excel(filepath)

    df = pd.read_excel(filepath, sheet_name=SHEET_NAME)
    df["Expiry Date"] = pd.to_datetime(df["Expiry Date"]).dt.strftime("%Y-%m-%d")
    df["Added Date"] = pd.to_datetime(df["Added Date"]).dt.strftime("%Y-%m-%d")
    df = df.fillna("")
    return df.to_dict(orient="records")


def save_data(items, filepath=EXCEL_FILE):
    if not items:
        df = pd.DataFrame(columns=COLUMNS)
    else:
        df = pd.DataFrame(items, columns=COLUMNS)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    header_fill = PatternFill("solid", start_color="1B4332")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    alt_fill = PatternFill("solid", start_color="F0FFF4")
    expiry_warn_fill = PatternFill("solid", start_color="FFF3CD")
    expiry_danger_fill = PatternFill("solid", start_color="FFDDD6")
    border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    today = date.today()
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        row_data = list(row)
        try:
            expiry = datetime.strptime(str(row_data[5]), "%Y-%m-%d").date()
            days_left = (expiry - today).days
        except:
            days_left = 999

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(vertical="center")
            cell.border = border
            if days_left <= 1:
                cell.fill = expiry_danger_fill
            elif days_left <= 3:
                cell.fill = expiry_warn_fill
            elif row_idx % 2 == 0:
                cell.fill = alt_fill

    col_widths = [6, 22, 15, 10, 10, 14, 14, 25]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    wb.save(filepath)
    return True


def get_next_id(items):
    if not items:
        return 1
    return max(item.get("ID", 0) for item in items) + 1


def get_expiring_soon(items, days=3):
    today = date.today()
    expiring = []
    for item in items:
        try:
            expiry = datetime.strptime(str(item.get("Expiry Date", "")), "%Y-%m-%d").date()
            days_left = (expiry - today).days
            if days_left <= days:
                item_copy = dict(item)
                item_copy["days_left"] = days_left
                expiring.append(item_copy)
        except:
            pass
    return sorted(expiring, key=lambda x: x["days_left"])


if __name__ == "__main__":
    cmd = sys.argv[1] if len(sys.argv) > 1 else "load"

    if cmd == "load":
        items = load_data()
        print(json.dumps({"status": "ok", "items": items, "categories": CATEGORIES}))

    elif cmd == "save":
        data = json.loads(sys.argv[2])
        save_data(data)
        print(json.dumps({"status": "ok"}))

    elif cmd == "expiring":
        items = load_data()
        days = int(sys.argv[2]) if len(sys.argv) > 2 else 3
        expiring = get_expiring_soon(items, days)
        print(json.dumps({"status": "ok", "expiring": expiring}))

    elif cmd == "create":
        create_excel()
        print(json.dumps({"status": "ok", "message": "Sample Excel file created"}))
